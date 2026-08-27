"""
import_osm_report.py -- OSM Excel → HRIS Dashboard

Usage:
    python import_osm_report.py
    python import_osm_report.py "C:\\path\\to\\All Open Tasks by Team.xlsx"
    python import_osm_report.py --file "<path>" --yes

Reads the most recently modified "All Open Tasks by Team*" report
(.xlsx or legacy .xls) from C:\\Users\\admin\\Downloads and pushes
data/tickets.json to GitHub. The v2 dashboard (index.html) is a static
file -- it fetches tickets.json dynamically and must not be overwritten
by this script.

OSM Administration's export format is not consistent -- some cycles
arrive as modern .xlsx, others as genuine old-format .xls (OLE binary,
not just a renamed extension). Both are handled natively here: no
manual conversion step, no LibreOffice dependency. .xlsx is read with
openpyxl; .xls is read with xlrd. xlrd is imported LAZILY (only when an
.xls file is actually selected) so an .xlsx-only run never fails just
because xlrd happens to be missing.

Requires:
    pip install openpyxl requests          (xlrd only needed for legacy .xls)
    Windows User env var: GITHUB_PAT (fine-grained PAT, Contents RW on hris-dashboard)

Automation hooks (used by the Downloads watcher and the morning Scheduled
Task; a human running this by hand normally needs none of them):
    OSM_IMPORT_FILE=<path>        import exactly this file, skip the newest-match search
    OSM_IMPORT_NONINTERACTIVE=1   never prompt; print warnings and proceed
                                  (also auto-detected when stdin is not a TTY)
    --file <path> / --yes / -y    same as the two env vars above

Every run prints a timestamp on its first line (estate-wide requirement).
"""

import os
import sys
import glob
import base64
import json
import re
import time
import zipfile
import tempfile
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

try:
    import requests
except ImportError:
    sys.exit("ERROR: requests not installed. Run: pip install requests")

# ── Config ────────────────────────────────────────────────────────────────────────────────────

DOWNLOADS     = r"C:\Users\admin\Downloads"
FILE_PATTERNS = [
    "All Open Tasks by Team*.xlsx",
    "All Open Tasks by Team*.xls",
]
GITHUB_REPO  = "begb0037admin/hris-dashboard"
GITHUB_REF   = "main"
BASE_API     = f"https://api.github.com/repos/{GITHUB_REPO}/contents"
DATA_API     = f"{BASE_API}/data/tickets.json"

# Column indices (0-based) -- fallback if header detection fails
COL_OWNER_TEAM  = 0
COL_TASK_NUMBER = 5
COL_SUMMARY     = 6
COL_STATUS      = 7
COL_PRIORITY    = 8
COL_ANALYST     = 9
COL_CATEGORY    = 4
COL_CREATED     = 11
COL_DAYS_OPEN   = 15

# Display order on the dashboard
KNOWN_ANALYSTS = [
    "James Salas Guillen",
    "Asta Palmer",
    "Michael O'Sullivan",
    "Simon Burford",
    "Kevin Lelitte",
]

STALE_THRESHOLD_DAYS = 30

# Robustness knobs (hardening added 27 Aug 2026, Drew -- see RESUME.md)
STALE_FILE_MINUTES = 20      # warn/prompt if the newest matching export is older than this
READ_RETRIES       = 3       # retries when the file looks locked / still downloading
READ_RETRY_WAIT    = 4       # seconds between those retries
LOCK_PATH          = os.path.join(tempfile.gettempdir(), "hris_osm_import.lock")
LOCK_STALE_SECONDS = 900     # a lock file older than this is ignored (assumed abandoned)
LOCK_WAIT_SECONDS  = 90      # how long to wait for another import to finish before proceeding anyway

# ── Interactive / automation mode ────────────────────────────────────────────

def _argv_flag(*names):
    return any(n in sys.argv[1:] for n in names)

def is_noninteractive():
    if str(os.environ.get("OSM_IMPORT_NONINTERACTIVE", "")).strip().lower() in ("1", "true", "yes"):
        return True
    if _argv_flag("--yes", "-y"):
        return True
    try:
        return not sys.stdin.isatty()
    except Exception:
        return True

def _requested_file():
    """Explicit file via --file <path>, OSM_IMPORT_FILE, or a bare argv[1] path."""
    env = os.environ.get("OSM_IMPORT_FILE", "").strip().strip('"')
    if env:
        return env
    argv = sys.argv[1:]
    for i, a in enumerate(argv):
        if a == "--file" and i + 1 < len(argv):
            return argv[i + 1].strip().strip('"')
    for a in argv:
        if not a.startswith("-") and a.lower().endswith((".xls", ".xlsx")):
            return a.strip().strip('"')
    return None

# ── Step 1: find the report file ────────────────────────────────────────────

_JUNK_PREFIXES = ("~$", ".~lock.")
_JUNK_SUFFIXES = (".crdownload", ".part", ".tmp", ".partial")

def _looks_like_junk(path):
    b = os.path.basename(path).lower()
    return b.startswith(_JUNK_PREFIXES) or b.endswith(_JUNK_SUFFIXES)

def _describe(path):
    mtime = datetime.fromtimestamp(os.path.getmtime(path))
    age_min = (datetime.now() - mtime).total_seconds() / 60.0
    print(f"Found report: {os.path.basename(path)}")
    print(f"  Full path:  {path}")
    print(f"  Modified:   {mtime.strftime('%d %b %Y %H:%M')}  ({age_min:.0f} min ago)")
    return age_min

def _stale_gate(age_min):
    if age_min <= STALE_FILE_MINUTES:
        return
    msg = (f"  NOTE: this is the newest matching export but it is {age_min:.0f} minutes old.\n"
           f"        If you just saved a fresh export it may not have finished downloading yet.")
    if is_noninteractive():
        print(msg)
        print("        (non-interactive mode -- proceeding with this file)")
        return
    print(msg)
    ans = input("  Continue with this file anyway? [y/N] ").strip().lower()
    if ans not in ("y", "yes"):
        sys.exit("Aborted -- no import performed. Re-run once the fresh export has downloaded.")

def find_report():
    explicit = _requested_file()
    if explicit:
        if not os.path.isfile(explicit):
            sys.exit(f"ERROR: requested file does not exist: {explicit}")
        if not explicit.lower().endswith((".xls", ".xlsx")):
            sys.exit(f"ERROR: requested file is not .xls/.xlsx: {explicit}")
        age_min = _describe(explicit)
        _stale_gate(age_min)
        return explicit

    files = []
    for pattern in FILE_PATTERNS:
        files.extend(glob.glob(os.path.join(DOWNLOADS, pattern)))
    files = [f for f in files if not _looks_like_junk(f)]
    if not files:
        patterns_str = " / ".join(FILE_PATTERNS)
        sys.exit(f"ERROR: No file matching '{patterns_str}' found in {DOWNLOADS}")
    latest = max(files, key=os.path.getmtime)
    age_min = _describe(latest)
    _stale_gate(age_min)
    return latest

# ── Step 2: read raw rows, format-agnostic ──────────────────────────────────
#
# Both readers normalise to the same shape openpyxl's
# `ws.iter_rows(values_only=True)` produces: a list of row-tuples, each
# cell either None, a str, an int/float, or a datetime. Everything
# downstream (header detection, column mapping, per-cell parsing) is
# then format-agnostic and untouched by which reader was used.

def _read_rows_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))

def _read_rows_xls(path):
    try:
        import xlrd
    except ImportError:
        sys.exit(
            "ERROR: this is a legacy .xls file and xlrd is not installed.\n"
            "Run: pip install xlrd   (xlrd is only needed for old-format .xls exports)"
        )
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    rows = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            ctype = cell.ctype
            value = cell.value
            if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                value = None
            elif ctype == xlrd.XL_CELL_DATE:
                try:
                    value = xlrd.xldate_as_datetime(value, wb.datemode)
                except (ValueError, OverflowError):
                    pass  # leave raw serial if it isn't a valid date
            elif ctype == xlrd.XL_CELL_NUMBER:
                if value == int(value):
                    value = int(value)
            elif ctype == xlrd.XL_CELL_BOOLEAN:
                value = bool(value)
            elif ctype == xlrd.XL_CELL_ERROR:
                value = None
            row.append(value)
        rows.append(tuple(row))
    return rows

def read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return _read_rows_xlsx(path)
    if ext == ".xls":
        return _read_rows_xls(path)
    sys.exit(f"ERROR: Unsupported file extension '{ext}' for {os.path.basename(path)}")

def _is_transient_read_error(e):
    """A file that is still downloading or open in Excel -- worth a retry.
    A genuinely malformed/schema-wrong file is not: let that crash loudly."""
    if isinstance(e, (PermissionError, zipfile.BadZipFile)):
        return True
    if type(e).__name__ in ("BadZipFile", "XLRDError", "CompDocError"):
        return True
    text = str(e).lower()
    if "being used by another process" in text or "permission denied" in text or "cannot access" in text:
        return True
    if isinstance(e, OSError):
        return True
    return False

def read_rows_with_retry(path):
    for attempt in range(1, READ_RETRIES + 1):
        try:
            return read_rows(path)
        except Exception as e:  # noqa: BLE001 -- narrowed immediately below
            if not _is_transient_read_error(e) or attempt == READ_RETRIES:
                if _is_transient_read_error(e):
                    sys.exit(
                        "ERROR: the export file could not be read after "
                        f"{READ_RETRIES} attempts.\n"
                        "It looks like it is still downloading, or still open in Excel.\n"
                        "Close it in Excel, wait for the download to finish, then run this again."
                    )
                raise
            print(
                f"  The export file appears to still be downloading, or is open in Excel "
                f"(attempt {attempt}/{READ_RETRIES}): {type(e).__name__}"
            )
            print(f"  Retrying in {READ_RETRY_WAIT}s -- close it in Excel if it is open...")
            time.sleep(READ_RETRY_WAIT)

# ── Step 3: parse the rows into tickets ───────────────────────────────

def parse_report(path):
    rows = read_rows_with_retry(path)

    header_idx = None
    for i, row in enumerate(rows):
        if any(str(c).strip() == "Task Number" for c in row if c is not None):
            header_idx = i
            break

    if header_idx is None:
        sys.exit("ERROR: Could not find header row containing 'Task Number'.")

    print(f"  Header row: {header_idx + 1} (1-indexed)")

    header = rows[header_idx]
    col_map = {str(cell).strip(): idx for idx, cell in enumerate(header) if cell is not None}

    def col(name, fallback):
        return col_map.get(name, fallback)

    c_task     = col("Task Number", COL_TASK_NUMBER)
    c_summary  = col("Summary",     COL_SUMMARY)
    c_status   = col("Status",      COL_STATUS)
    c_priority = col("Priority",    COL_PRIORITY)
    c_analyst  = col("Analyst",     COL_ANALYST)
    c_category = col("Category",    COL_CATEGORY)
    c_created  = col("Created",     COL_CREATED)
    c_days     = col("Days Open",   COL_DAYS_OPEN)
    c_team     = col("Owner Team",  COL_OWNER_TEAM)

    tickets = []
    for row in rows[header_idx + 1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(idx):
            try:
                v = row[idx]
                return str(v).strip() if v is not None else ""
            except IndexError:
                return ""

        task_num = cell(c_task)
        if not task_num:
            continue

        created = cell(c_created)
        raw_created = row[c_created] if c_created < len(row) else None
        if isinstance(raw_created, datetime):
            created = raw_created.strftime("%d %b %Y")
        elif created and re.match(r"\d{4}-\d{2}-\d{2}", created):
            try:
                created = datetime.strptime(created[:10], "%Y-%m-%d").strftime("%d %b %Y")
            except ValueError:
                pass

        days_raw = cell(c_days)
        try:
            days_int = int(float(days_raw)) if days_raw else 0
        except ValueError:
            days_int = 0

        tickets.append({
            "task_num": task_num,
            "summary":  cell(c_summary),
            "status":   cell(c_status),
            "priority": cell(c_priority),
            "analyst":  cell(c_analyst),
            "category": cell(c_category),
            "created":  created,
            "days":     days_int,
            "team":     cell(c_team),
        })

    print(f"  Tickets parsed: {len(tickets)}")
    return tickets

# ── Step 4: group tickets ─────────────────────────────────────────────────────

def group_tickets(tickets):
    by_analyst = {name: [] for name in KNOWN_ANALYSTS}
    unassigned = []
    other = {}

    for t in tickets:
        a = t["analyst"].strip()
        if not a:
            unassigned.append(t)
        elif a in by_analyst:
            by_analyst[a].append(t)
        else:
            other.setdefault(a, []).append(t)

    for name in sorted(other):
        by_analyst[name] = other[name]

    return by_analyst, unassigned

# ── Step 5: build tickets.json ────────────────────────────────────────────────────────────

def build_tickets_json(by_analyst, unassigned, report_path):
    all_tickets = [t for ts in by_analyst.values() for t in ts] + unassigned
    total     = len(all_tickets)
    assigned  = total - len(unassigned)
    stale     = [t for t in all_tickets if t["days"] >= STALE_THRESHOLD_DAYS]
    oldest    = max((t["days"] for t in all_tickets), default=0)

    now = datetime.now()

    analysts_list = []
    for name in list(by_analyst.keys()):
        tickets = by_analyst[name]
        if tickets or name in KNOWN_ANALYSTS:
            analysts_list.append({
                "name": name,
                "ticket_count": len(tickets),
                "tickets": tickets,
            })

    return {
        "updated":         now.strftime("%Y-%m-%dT%H:%M:%S"),
        "updated_display": now.strftime("%A %d %B %Y at %H:%M"),
        "report_file":     os.path.basename(report_path),
        "summary": {
            "total":      total,
            "assigned":   assigned,
            "unassigned": len(unassigned),
            "stale":      len(stale),
            "oldest_days": oldest,
        },
        "analysts":   analysts_list,
        "unassigned": unassigned,
    }

# ── Step 6: GitHub push helpers ──────────────────────────────────────────────────────────

def get_github_pat():
    pat = os.environ.get("GITHUB_PAT", "").strip()
    if not pat:
        sys.exit(
            "ERROR: GITHUB_PAT environment variable is not set.\n"
            "Set it as a Windows User environment variable (never in any file)."
        )
    return pat

def github_put(api_url, content_bytes, commit_message, pat):
    """GET current SHA, then PUT new content. Returns new SHA.

    If the PUT loses a race (another import pushed tickets.json between our
    GET and PUT -- HTTP 409/422 'sha' conflict), re-GET the current SHA and
    retry the PUT once. Concurrent imports then self-heal: whichever runs
    last wins and tickets.json is always a complete, valid single import.
    """
    headers = {
        "Authorization": f"token {pat}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    def _get_sha():
        r = requests.get(f"{api_url}?ref={GITHUB_REF}", headers=headers, timeout=30)
        if r.status_code == 404:
            return None
        if r.status_code == 200:
            return r.json()["sha"]
        sys.exit(f"ERROR: GET {api_url} failed (HTTP {r.status_code})\n{r.text}")

    def _put(current_sha):
        payload = {
            "message": commit_message,
            "content": base64.b64encode(content_bytes).decode("ascii"),
            "branch":  GITHUB_REF,
        }
        if current_sha:
            payload["sha"] = current_sha
        return requests.put(api_url, headers=headers, json=payload, timeout=30)

    current_sha = _get_sha()
    put_resp = _put(current_sha)

    if put_resp.status_code in (409, 422):
        print(f"  Push race detected (HTTP {put_resp.status_code}) -- re-reading and retrying once...")
        time.sleep(2)
        put_resp = _put(_get_sha())

    if put_resp.status_code not in (200, 201):
        sys.exit(f"ERROR: PUT {api_url} failed (HTTP {put_resp.status_code})\n{put_resp.text}")

    return put_resp.json()["content"]["sha"]

# ── Advisory single-writer lock ──────────────────────────────────────────────
#
# Best-effort only. Every path that updates tickets.json runs this script
# (manual .bat, morning Scheduled Task, Downloads watcher), so one lock here
# covers all three. It can never deadlock: if another import holds a fresh
# lock we wait at most LOCK_WAIT_SECONDS and then proceed anyway (the
# github_put race-retry above is the real safety net).

def acquire_lock():
    try:
        if os.path.exists(LOCK_PATH):
            age = time.time() - os.path.getmtime(LOCK_PATH)
            if age < LOCK_STALE_SECONDS:
                waited = 0
                print(f"  Another import appears to be running (lock {age:.0f}s old) -- waiting up to {LOCK_WAIT_SECONDS}s...")
                while os.path.exists(LOCK_PATH) and waited < LOCK_WAIT_SECONDS:
                    time.sleep(5)
                    waited += 5
                    try:
                        if time.time() - os.path.getmtime(LOCK_PATH) >= LOCK_STALE_SECONDS:
                            break
                    except OSError:
                        break
                if os.path.exists(LOCK_PATH):
                    print("  Lock still present -- proceeding anyway (push uses a race-safe retry).")
        with open(LOCK_PATH, "w", encoding="utf-8") as fh:
            fh.write(f"pid={os.getpid()} at {datetime.now().isoformat(timespec='seconds')}\n")
    except OSError as e:
        print(f"  (could not create lock file, continuing without it: {e})")

def release_lock():
    try:
        if os.path.exists(LOCK_PATH):
            os.remove(LOCK_PATH)
    except OSError:
        pass

# ── Main ────────────────────────────────────────────────────────────────────────────────────

def main():
    print(f"=== import_osm_report.py -- {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
    if is_noninteractive():
        print("    (non-interactive mode)")
    print()

    acquire_lock()
    try:
        report_path            = find_report()
        tickets                = parse_report(report_path)
        by_analyst, unassigned = group_tickets(tickets)

        total = sum(len(v) for v in by_analyst.values()) + len(unassigned)
        print()
        print("Ticket summary:")
        for name, t in by_analyst.items():
            if t:
                print(f"  {name}: {len(t)}")
        print(f"  Unassigned: {len(unassigned)}")
        print(f"  Total: {total}")
        print()

        pat     = get_github_pat()
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        print("Pushing data/tickets.json...")
        payload    = build_tickets_json(by_analyst, unassigned, report_path)
        json_bytes = json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")
        new_sha    = github_put(DATA_API, json_bytes, f"OSM import — {now_str}", pat)
        print(f"  data/tickets.json SHA: {new_sha}")
        print("  Done.")
        print()
    finally:
        release_lock()

    print(f"All updates pushed successfully at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.")
    print("GitHub Pages usually takes 60-90 seconds to rebuild after a push.")
    print("If the dashboard still shows the old time, wait a minute and hard-reload (Ctrl-Shift-R):")
    print("  https://begb0037admin.github.io/hris-dashboard/")

if __name__ == "__main__":
    main()
